from __future__ import division
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string

resultBook = openpyxl.load_workbook('resultSet.xlsx')
ws = resultBook.get_sheet_by_name("results")

resultBook.create_sheet('sorted')
ss = resultBook.get_sheet_by_name('sorted')

rows = ws.max_row
cols = ws.max_column

header = [cell.value for cell in ws[1]]
data = []
data_dict = {}
sorted_list = []

for row in ws[2:rows]:
    data_row = []
    for cell in row:
        data_row.append(cell.value)
    index = data_row[cols - 1]
    data_dict[int(index)] = data_row
    data.append(data_row)

# sorting
keylist = data_dict.keys()
keylist.sort()
for key in keylist:
    sorted_list.append(data_dict[key])

d_line = 2
for row in ss['A' + str(d_line - 1):'AR' + str(d_line - 1)]:
    col_index = 0
    for cell in row:
        cell.value = header[col_index]
        col_index = col_index + 1
for data_row in sorted_list:
    for row in ss['A' + str(d_line):'AR' + str(d_line)]:
        col_index = 0
        for cell in row:
            cell.value = data_row[col_index]
            col_index = col_index + 1
    d_line = d_line + 1

resultBook.save('sorted.xlsx')
